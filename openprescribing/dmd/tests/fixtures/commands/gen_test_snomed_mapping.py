# This script takes a SNOMED to BNF code mapping spreadsheet, and generates a
# new spreadsheet with the subset of the rows corresponding to given VMPs.  The
# script relies having acces to a databse against which import_dmd has been
# run.
#
# Usage:
#
#   python gen_test_snomed_mapping.py [inp_path] [outp_path] [vpids]
#
# Where:
#
#   inp_path is the path of the input spreadsheet
#   outp_path is the path of the new spreadsheet
#   vpids is a comma-separated list of IDs of VMPs



import os
import sys

from openpyxl import Workbook, load_workbook
import psycopg2

if len(sys.argv) != 4:
    print('Usage: python gen_test_snomed_mapping.py [inp_path] [outp_path] [vpids]')
    sys.exit(1)

inp_path = sys.argv[1]
outp_path = sys.argv[2]
vpids = sys.argv[3].split(',')

connection = psycopg2.connect(
    database=os.getenv('DB_NAME'),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASS'),
)

cursor = connection.cursor()

snomed_codes = []

for vpid in vpids:
    cursor.execute('SELECT vppid FROM dmd_vmpp WHERE vpid = %s', [vpid])
    for row in cursor.fetchall():
        vppid = row[0]
        snomed_codes.append(vppid)

        cursor.execute('SELECT appid FROM dmd_ampp WHERE vppid = %s', [vppid])
        for row in cursor.fetchall():
            snomed_codes.append(row[0])

assert len(snomed_codes) == len(set(snomed_codes))
snomed_codes = set(snomed_codes)

wb_in = load_workbook(inp_path)
wb_out = Workbook()

rows = wb_in.active.rows

headers = rows[0]
assert headers[0].value == 'BNF Code'
assert headers[2].value == 'VMPP / AMPP SNOMED Code'

ws = wb_out.active
header_values = [cell.value for cell in rows[0]]
ws.append(header_values)

for row in rows[1:]:
    snomed_code = row[2].value

    if not snomed_code:
        continue

    if snomed_code[0] == "'":
        snomed_code = snomed_code[1:]
        if not snomed_code:
            continue

    if int(snomed_code) in snomed_codes:
        row_values = [cell.value for cell in row]
        print(snomed_code)
        ws.append(row_values)

wb_out.save(outp_path)
